import openpyxl
import json
import glob

month = "03"
year = "2018"
nd = 0
breakfast_start = 5
breakfast_end = 13
lunch_start = 16
lunch_end = 24
dinner_start = 27
dinner_end = 34
no_of_days = 15

for wb in glob.glob("*.xlsx"):
    messywb = openpyxl.load_workbook(wb)
    messysheet = messywb.get_active_sheet()

    for i in range(1, no_of_days + 1):
        breakfast = []
        lunch = []
        dinner = []
        if i < 10:
            date = "0" + str(i)
        else:
            date = str(i)
        filename = "../db/menu/" + date + "-" + month + "-" + year + ".json"
        for j in range(breakfast_start, breakfast_end + 1):
            # add to breakfast if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                breakfast.append(temp.strip())
        for j in range(lunch_start, lunch_end + 1):
            # add to lunch if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                lunch.append(temp.strip())
        for j in range(dinner_start, dinner_end + 1):
            # add to dinner if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                dinner.append(temp.strip())

        menu = {
            "breakfast": breakfast,
            "lunch": lunch,
            "dinner": dinner
        }
        json_string = json.dumps(menu, indent=4)
        fi = open(filename, "w+")
        fi.write(json_string)
        fi.close()
