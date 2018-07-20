# edit the argument file for the extraction only one file at a time should
# be in this directory
import openpyxl
import json
import glob

month = "03"
year = "2018"

for wb in glob.glob("*.xlsx"):
    messywb = openpyxl.load_workbook(wb)
    messysheet = messywb.get_active_sheet()
    nd = 0
    with open("arg.json") as json_data:
        argu = json.load(json_data)
        json_data.close()
        breakfast_start = bs = int(argu["breakfast_start"])
        breakfast_end = be = int(argu["breakfast_end"])
        lunch_start = ls = int(argu["lunch_start"])
        lunch_end = le = int(argu["lunch_end"])
        dinner_start = ds = int(argu["dinner_start"])
        dinner_end = de = int(argu["dinner_end"])
        no_of_days = nd = int(argu["no_of_days"])

    for i in range(1, nd + 1):
        breakfast = []
        lunch = []
        dinner = []
        if i < 10:
            date = "0" + str(i)
        else:
            date = str(i)
        filename = "../db/menu/" + date + "-" + month + "-" + year + ".json"
        for j in range(bs, be + 1):
            # add to breakfast if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                breakfast.append(temp.strip())
        for j in range(ls, le + 1):
            # add to lunch if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                lunch.append(temp.strip())
        for j in range(ds, de + 1):
            # add to dinner if not empty
            temp = messysheet.cell(row=j, column=i).value
            if(temp is not None and temp[0].isalpha()):
                dinner.append(temp.strip())
        menu = {
            "breakfast": breakfast,
            "lunch": lunch,
            "dinner": dinner
        }
        json_string = json.dumps(menu, indent=2)
        fi = open(filename, "w+")
        fi.write(json_string)
        fi.close()
